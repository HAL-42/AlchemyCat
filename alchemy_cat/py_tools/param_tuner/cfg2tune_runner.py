#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
@Author  : Xiaobo Yang
@Contact : hal_42@zju.edu.cn
@Time    : 2021/5/27 14:08
@File    : cfg2tune_runner.py
@Software: PyCharm
@Desc    : 
"""
import os.path as osp
import subprocess
import sys
import traceback
from functools import wraps
from multiprocessing import Pool
from pprint import pprint
from typing import List, Optional, Any, Callable

if sys.version_info >= (3, 11):
    from typing import TypeAlias
else:  # 兼容Python<3.11。
    TypeAlias = Any

import pandas as pd
from openpyxl import load_workbook
from tqdm import tqdm

from .cfg2tune import Cfg2Tune
from ..config.py_cfg import Config
from ..logger import Logger
from ..str_formatters import get_local_time_str
from ...cuda_tools import allocate_cuda_by_group_rank, get_cudas
from ..color_print import yprint, gprint, rprint

__all__ = ["Cfg2TuneRunner"]

WORK_WRAPPER_INPUT_TYPE: TypeAlias = tuple[int, tuple[Config, str, str, dict[str, Any]]]
WORK_INPUT_TYPE: TypeAlias = tuple[int, Config, str, str, dict[str, str]]
WORK_WRAPPER_TYPE: TypeAlias = Callable[[WORK_WRAPPER_INPUT_TYPE], Any]
WORK_TYPE: TypeAlias = Callable[[int, Config, str, str, dict[str, str]], Any]


class Cfg2TuneRunner(object):
    """Running cfg2tune"""
    def __init__(self, cfg2tune_py: str, config_root: str='./configs', experiment_root="experiment",
                 pool_size: int=0, work_gpu_num: int=None,
                 metric_names: Optional[List[str]]=None,
                 gather_metric_fn: Callable[[Config, str, Any, dict[str, tuple[Any, str]]], dict[str, Any]]=None,
                 work_fn: WORK_WRAPPER_TYPE=None,
                 allow_fail_gather: bool=True,
                 block: bool=True, verbosity: bool=True,
                 memory_need: float=-1., max_process: int=-1):
        """Running a cfg2tune.

        Args:
            cfg2tune_py: cfg2tune python script.
            config_root: Root dir where configs store.
            experiment_root: Root dir where experiment results store.
            pool_size: Multiprocess' Pool size when running tuning work.
            work_gpu_num: Number of GPUs in a group.
            metric_names: Name of metrics to be gathered.
            allow_fail_gather: Whether allow gather_metric_fn to fail.
            block: Refer to `alchemy_cat.cuda_tools.allocate_cuda_by_group_rank`.
            verbosity: Refer to `alchemy_cat.cuda_tools.allocate_cuda_by_group_rank`.
            memory_need: Refer to `alchemy_cat.cuda_tools.allocate_cuda_by_group_rank`.
            max_process: Refer to `alchemy_cat.cuda_tools.allocate_cuda_by_group_rank`.
        """
        self.cfg2tune_py = cfg2tune_py
        self.config_root = config_root
        self.experiment_root = experiment_root
        self.gather_metric_fn = gather_metric_fn
        self.work_fn = work_fn
        self.work_gpu_num = int(work_gpu_num) if work_gpu_num is not None else work_gpu_num
        self.allow_fail_gather = allow_fail_gather

        # -* 解算pool_size。
        if self.work_gpu_num is not None:
            if work_gpu_num <= 0:
                raise ValueError("work_gpu_num must be None or integer larger than 0.")
            if pool_size != 0:
                raise ValueError("work_gpu_num is set, pool size will be calculated. ")
            self.pool_size = len(get_cudas()[0]) // self.work_gpu_num
        else:
            self.pool_size = int(pool_size)

        # * 加载Cfg2Tune。
        self.cfg2tune = Cfg2Tune.load_cfg2tune(cfg2tune_py, experiments_root=experiment_root, config_root=config_root)

        # * 设置整个调参实验的结果文件夹。
        self.rslt_dir = self.cfg2tune.rslt_dir

        # * 获取指标名。
        if self.cfg2tune.metric_names:
            assert metric_names is None
            self.metric_names = self.cfg2tune.metric_names
        else:
            self.metric_names = metric_names

        # * 所有参数组合+参数组合对应的metric。
        self.param_combs: List[dict[str, tuple[Any, str]]] = self.cfg2tune.param_combs
        self.metric_frame: Optional[pd.DataFrame] = None

        # * 每个配置的pkl及其对应的实验文件夹。
        self.cfg_pkls: List[str] = []
        self.cfgs: list[Config] = []
        self.cfg_rslt_dirs: List[str] = []

        # * 保存每个配置的运行结果。
        self.run_rslts: List[subprocess.CompletedProcess] = []
        self.cuda_env_kwargs = {'group_num': self.pool_size,
                                'block': block, 'verbosity': verbosity,
                                'memory_need': memory_need, 'max_process': max_process}

    def build_metrics(self):
        """得到参数组合与metric frame。"""
        midx = pd.MultiIndex.from_tuples([tuple(val[1] for val in param_comb.values())
                                          for param_comb in self.param_combs],
                                         names=list(self.param_combs[0].keys()))
        self.metric_frame = pd.DataFrame(index=midx, columns=self.metric_names)

    def set_cfgs(self):
        """将Cfg2Tune转存为pkl文件，并得到对应的实验文件夹。"""
        self.cfg_pkls, self.cfgs = self.cfg2tune.dump_cfgs(self.config_root)
        self.cfg_rslt_dirs = [cfg['rslt_dir'] for cfg in self.cfgs]

    def run_cfgs(self):
        """并行或串行地，根据每个配置，执行work函数。work函数内，应当完成一次实验。"""
        work = self.work_fn if self.work_fn is not None else self.work

        work_inputs = enumerate(zip(self.cfgs, self.cfg_pkls, self.cfg_rslt_dirs,
                                    [self.cuda_env_kwargs] * len(self.cfgs)))

        if self.pool_size > 0:
            with Pool(self.pool_size) as p:
                map_it = p.imap(work, work_inputs, chunksize=1)
                for run_rslt in tqdm(map_it, 'Tuning', len(self.cfg_pkls), unit='configs', dynamic_ncols=True):
                    self.run_rslts.append(run_rslt)
        else:
            for work_input in tqdm(work_inputs, 'Tuning', len(self.cfg_pkls), unit='configs', dynamic_ncols=True):
                try:
                    run_rslt = work(work_input)
                except subprocess.CalledProcessError as e:
                    print(f"Cfg2TuneRunner's work failed. The stdout and stderr of work are: ")
                    print(f"{e.stdout}\n\n{e.stderr}")
                    raise e
                self.run_rslts.append(run_rslt)

    @staticmethod
    def parse_work_param(pkl_idx_cfg_cfg_pkl_cfg_rslt_dir_cuda_env_kwargs: WORK_WRAPPER_INPUT_TYPE) -> WORK_INPUT_TYPE:
        """Convert `WORK_WRAPPER_INPUT_TYPE` to `WORK_INPUT_TYPE`"""
        pkl_idx, (cfg, cfg_pkl, cfg_rslt_dir, cuda_env_kwargs) = pkl_idx_cfg_cfg_pkl_cfg_rslt_dir_cuda_env_kwargs
        return pkl_idx, cfg, cfg_pkl, cfg_rslt_dir, allocate_cuda_by_group_rank(pkl_idx, **cuda_env_kwargs)[1]

    @staticmethod
    def work(pkl_idx_cfg_cfg_pkl_cfg_rslt_dir_cuda_env_kwargs: WORK_WRAPPER_INPUT_TYPE):
        """Do not implement this function, use `register_work_fn` instead."""
        raise NotImplementedError()

    def register_work_fn(self, work_fn: WORK_TYPE) -> WORK_WRAPPER_TYPE:
        """Register work_fn. Use as a decorator."""
        @wraps(work_fn)
        def wrapper(pkl_idx_cfg_cfg_pkl_cfg_rslt_dir_cuda_env_kwargs: WORK_WRAPPER_INPUT_TYPE):
            pkl_idx, cfg, cfg_pkl, cfg_rslt_dir, cuda_env = (
                Cfg2TuneRunner.parse_work_param(pkl_idx_cfg_cfg_pkl_cfg_rslt_dir_cuda_env_kwargs))
            return work_fn(pkl_idx, cfg, cfg_pkl, cfg_rslt_dir, cuda_env)

        self.work_fn = wrapper

        return wrapper

    def gather_metrics(self):
        gather_failed = False
        for cfg, cfg_rslt_dir, run_rslt, param_comb in zip(self.cfgs, self.cfg_rslt_dirs,
                                                           self.run_rslts, self.param_combs):
            try:
                metric = self.gather_metric(cfg, cfg_rslt_dir, run_rslt, param_comb)
            except Exception as e:
                if self.allow_fail_gather:
                    yprint(f"[WARNING] Error occurred when gather_metric for {cfg_rslt_dir}. ")
                    rprint(traceback.format_exc())
                    gather_failed = True
                    continue
                else:
                    raise e

            if self.metric_frame is None:
                if self.metric_names is None:
                    self.metric_names = list(metric.keys())
                self.build_metrics()
            else:
                assert set(metric.keys()) == set(self.metric_names), \
                    f"metric.keys()={metric.keys()} != self.metric_names={self.metric_names}"

            self.metric_frame.loc[tuple(val[1] for val in param_comb.values())] = metric

        if gather_failed:
            yprint("[WARNING] Some gather_metric failed. Please check the error message above.")

    def gather_metric(self, cfg: Config, cfg_rslt_dir: str, run_rslt: Any, param_comb: dict[str, tuple[Any, str]]) \
            -> dict[str, Any]:
        """Given cfg_rslt_dir, run_stdout, param_comb, return {metric_name: metric}"""
        if self.gather_metric_fn is not None:
            return self.gather_metric_fn(cfg, cfg_rslt_dir, run_rslt, param_comb)
        else:
            raise NotImplementedError("gather_metric not implemented")

    def register_gather_metric_fn(self,
                                  gather_metric_fn: Callable[[Config, str, Any, dict[str, tuple[Any, str]]],
                                                             dict[str, Any]]):
        """Register gather_metric_fn. Use as a decorator."""
        self.gather_metric_fn = gather_metric_fn
        return gather_metric_fn

    def save_metrics(self):
        exp_short_id = osp.basename(osp.dirname(self.cfg2tune_py)).split('@')[-1]
        # 美观起见，excel选择merge cells。需要时可以手动解除merge。
        self.metric_frame.to_excel(osp.join(self.rslt_dir, f'{exp_short_id}.xlsx'), sheet_name='MetricFrame',
                                   merge_cells=True)
        # 将刚才保存的xlsx转储为xlsm，用于数据分析。
        load_workbook(filename=osp.join(self.rslt_dir, f'{exp_short_id}.xlsx'),
                      read_only=False, keep_vba=True).save(osp.join(self.rslt_dir, f'{exp_short_id}.xlsm'))
        self.metric_frame.to_pickle(osp.join(self.rslt_dir, f'{exp_short_id}.pkl'))

    def tuning(self):
        stdout_log_file = osp.join(self.rslt_dir, 'stdout',
                                   '-'.join(['stdout', get_local_time_str(for_file_name=True)]) + '.log')
        Logger(stdout_log_file, real_time=True)
        print(f"Log stdout at {stdout_log_file}")

        gprint("-----------------Setting Configs-----------------")
        self.set_cfgs()

        print("Saving Config pkls at: ")
        pprint(self.cfg_pkls)

        print("Param Combines: ")
        pprint(self.param_combs, sort_dicts=False)

        gprint("-----------------Running Configs-----------------")
        self.run_cfgs()

        if self.gather_metric_fn is not None:
            gprint("-----------------Gather Metrics-----------------")
            self.gather_metrics()
            print("Metric Frame: ")
            pprint(self.metric_frame)

            gprint("-----------------Save Metrics-----------------")
            self.save_metrics()
            print(f"Saving Metric Frame at {osp.join(self.rslt_dir, 'metric_frame.xlsx')}")
